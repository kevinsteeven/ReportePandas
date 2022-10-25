from openpyxl import load_workbook
from openpyxl.chart import BarChart , Reference
wb=load_workbook('pivot_table.xlsx')
sheet=wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row=wb.active.max_row
barchart = BarChart()
#Definir referencias de los datos y las categorias
data = Reference(sheet, min_col = min_column + 1, min_row = min_row , max_col = max_column, max_row = max_row)
categories = Reference(sheet, min_col = min_column , min_row = min_row + 1, max_col = min_column, max_row = max_row)
#Añadir datos
barchart.add_data(data, titles_from_data=True)
#Definir categorias
barchart.set_categories(categories)
#Añadir grafico
sheet.add_chart(barchart,"B12")
barchart.title = "Ventas por linea de producto"
barchart.style = 2
wb.save('barchart.xlsx')
