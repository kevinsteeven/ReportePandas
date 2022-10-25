from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb=load_workbook('pivot_table.xlsx')
sheet=wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row=wb.active.max_row

for i in range(min_column+1,max_column +1):#Definir columnas en las que se aplicara la formula, la funcion range no cueta el valor final por eso se le suma uno a max_column
    letra_columna=get_column_letter(i)
    sheet[f'{letra_columna}{max_row +1}']=f'=SUM({letra_columna}{min_row + 1}:{letra_columna}{max_row})'#Añadir formula a celda
    sheet[f'{letra_columna}{max_row +1}'].style = 'Currency'



wb.save('Report.xlsx')