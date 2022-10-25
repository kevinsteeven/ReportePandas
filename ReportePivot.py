from calendar import month
from webbrowser import get
from openpyxl import load_workbook
from openpyxl.chart import BarChart , Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import os
import sys

app_path = os.path.dirname(sys.executable)


mes = input('Introduce el mes:')

input_path= os.path.join(app_path,'pivot_table.xlsx')

wb=load_workbook(input_path)
sheet=wb['Report']
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row=wb.active.max_row
barchart = BarChart()
#Definir referencias de los datos y las categorias
data = Reference(sheet, min_col = min_column + 1, min_row = min_row , max_col = max_column, max_row = max_row)
categories = Reference(sheet, min_col = min_column , min_row = min_row + 1, max_col = min_column, max_row = max_row)
#Añadir datos
barchart.add_data(data, titles_from_data=True)
#Definir categorias
barchart.set_categories(categories)
#Añadir grafico
sheet.add_chart(barchart,"B12")
barchart.title = "Ventas por linea de producto"
barchart.style = 2

for i in range(min_column+1,max_column +1):#Definir columnas en las que se aplicara la formula, la funcion range no cueta el valor final por eso se le suma uno a max_column
    letra_columna=get_column_letter(i)
    sheet[f'{letra_columna}{max_row +1}']=f'=SUM({letra_columna}{min_row + 1}:{letra_columna}{max_row})'#Añadir formula a celda
    sheet[f'{letra_columna}{max_row +1}'].style = 'Currency'

sheet[f'{get_column_letter(min_column)}{max_row+1}']='Total'
#Escribir titulo en celda A1
sheet['A1']='Reporte de ventas'
#Escribir subtitulo en celda A2
sheet['A2']=mes
#Editar fuentes
sheet['A1'].font=Font('Arial', bold= True, size=20)
sheet['A2'].font=Font('Arial', bold= True, size=10)

output_path= os.path.join(app_path,f'reporte_{mes}.xlsx')
wb.save(output_path)